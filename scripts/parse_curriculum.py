import os
import json
import logging
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from core import config

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

logger = config.get_logger("scripts.parse_curriculum")

def parse_curriculum():
    data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
    excel_path = os.path.join(data_dir, "Core & Electives.xlsx")
    json_path = os.path.join(data_dir, "curriculum.json")

    if not os.path.exists(excel_path):
        logger.error(f"Cannot find {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    curriculum = {}

    for sheet_name in ["Core", "Elective"]:
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in Excel.")
            continue
            
        ws = wb[sheet_name]
        current_semester = None
        current_program = None

        # Headers are usually on row 1: COURSE, TITLE, CREDITS, Course Type, Sections
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            
            c0 = str(row[0]).strip() if row[0] else ""
            
            # If column 0 starts with 'SEMESTER' it's a semester block
            if c0.upper().startswith("SEMESTER"):
                current_semester = c0.upper().replace("SEMESTER", "").strip()
                # Convert roman numerals if needed, but keeping it simple for now
                if current_semester == "I": current_semester = "1"
                elif current_semester == "II": current_semester = "2"
                elif current_semester == "III": current_semester = "3"
                elif current_semester == "IV": current_semester = "4"
                elif current_semester == "V": current_semester = "5"
                elif current_semester == "VI": current_semester = "6"
                elif current_semester == "VII": current_semester = "7"
                elif current_semester == "VIII": current_semester = "8"
                current_program = None
                continue
                
            # If column 1 is empty, it's a program block (e.g. 'B Tech (Institute Core)')
            if c0 and not row[1]:
                # Ignore sub-headers so they don't overwrite the main program
                if c0.strip().lower() in ['electives', 'open elective', 'core', 'technical elective', 'open electives']:
                    continue
                current_program = c0
                continue
                
            # Otherwise, it's a course row
            if c0 and row[1]:
                title = str(row[1]).strip()
                course_type = str(row[3]).strip() if len(row) > 3 and row[3] else sheet_name
                
                # Some courses may appear multiple times across programs/electives.
                # We will store lists if there are multiple, or just the first one. Let's just keep the latest for now,
                # or a list of programs if it's shared. For simplicity, we just store program and semester.
                # Actually, a course might be core for one program and elective for another.
                
                if c0 not in curriculum:
                    curriculum[c0] = []
                
                # Check if we already have this exact program+semester to avoid duplicates
                sem_val = int(current_semester) if current_semester and current_semester.isdigit() else current_semester
                exists = any(m.get("program") == current_program and m.get("semester") == sem_val for m in curriculum[c0])
                if not exists:
                    curriculum[c0].append({
                        "course_name": title,
                        "program": current_program,
                        "semester": sem_val,
                        "course_type": course_type
                    })

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(curriculum, f, indent=4)
        
    logger.info(f"Parsed {len(curriculum)} courses into curriculum.json")

if __name__ == "__main__":
    parse_curriculum()
