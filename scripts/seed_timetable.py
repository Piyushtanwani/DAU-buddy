"""
Seed Timetable Script
=====================
Parses the active lecture timetable Excel file (Lecture_TT_Autumn2026-27_v9.xlsx)
and populates the timetables table with clean program batch mappings using curriculum.json.

Usage:
    python scripts/seed_timetable.py
"""
import os
import re
import sys
import json
import psycopg2
import psycopg2.extras
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from core import config
from core.database import db_connection
from core.utils.venue import normalize_venue_id

logger = config.get_logger("scripts.seed_timetable")

try:
    import openpyxl
except ImportError:
    logger.error("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
POSSIBLE_PATHS = [
    os.path.join(DATA_DIR, "Lecture Data.xlsx"),
    os.path.join(DATA_DIR, "Lecture_TT_Autumn2026-27_v9.xlsx"),
]
POSSIBLE_LAB_PATHS = [
    os.path.join(DATA_DIR, "Lab Data.xlsx"),
    os.path.join(DATA_DIR, "Lab_TT_Autumn2026-27.xlsx"),
]

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

def find_excel_file() -> str:
    for path in POSSIBLE_PATHS:
        if os.path.exists(path):
            return path
    # Dynamic search for any Lecture*.xlsx in data/
    if os.path.exists(DATA_DIR):
        for f in os.listdir(DATA_DIR):
            if f.endswith(".xlsx") and "lecture" in f.lower():
                return os.path.join(DATA_DIR, f)
    raise FileNotFoundError(f"No timetable Excel file found in data/ folder ({DATA_DIR}).")

def find_lab_excel_file() -> str | None:
    for path in POSSIBLE_LAB_PATHS:
        if os.path.exists(path):
            return path
    if os.path.exists(DATA_DIR):
        for f in os.listdir(DATA_DIR):
            if f.endswith(".xlsx") and "lab" in f.lower():
                return os.path.join(DATA_DIR, f)
    return None

def load_curriculum():
    path = os.path.join(DATA_DIR, "curriculum.json")
    if not os.path.exists(path):
        logger.warning("curriculum.json not found. Run parse_curriculum.py first.")
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _parse_time(t) -> str | None:
    if not t:
        return None
    m = re.match(r"(\d{1,2}):(\d{2})", str(t))
    if not m:
        return None
    hour = int(m.group(1))
    # Timetable labels are 12-hour without AM/PM; campus day runs 08:00-19:00,
    # so hours 1-7 are afternoon (13:00-19:00).
    if 1 <= hour <= 7:
        hour += 12
    return f"{hour:02d}:{m.group(2)}:00"

def split_rooms(room_str: str) -> list:
    """Split compound room strings ("CEP-102 & CEP-110", "LAB207, LAB112 & LAB009")
    into normalized individual rooms. Returns [""] when no room is given so callers
    still emit one record."""
    parts = [p for p in re.split(r"\s*[&,]\s*", room_str or "") if p.strip()]
    return [normalize_venue_id(p) for p in parts] or [""]

def _resolve_slot(slot) -> tuple:
    if not slot:
        return None, None
    m = re.match(r"(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})", str(slot))
    if m:
        return _parse_time(m.group(1)), _parse_time(m.group(2))
    return None, None

def read_abbrev_map(wb) -> dict:
    """Short name → full display name, from the workbook's abbreviations sheet.

    e.g. {"AC": "Ankush Chander (AC)"}. Both parsers write this same display
    form into `timetables.faculty_name`, so lecture and lab rows for one person
    are matchable by the same name lookup.
    """
    abbrev_map = {}
    if "FacultyNameAbbreviations" in wb.sheetnames:
        for row in wb["FacultyNameAbbreviations"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                abbrev_map[str(row[1]).strip()] = str(row[0]).strip()
    return abbrev_map


def load_abbrev_map(excel_path: str) -> dict:
    """read_abbrev_map() for callers that only have a path (e.g. the lab parser,
    whose own workbook carries no abbreviations sheet)."""
    return read_abbrev_map(openpyxl.load_workbook(excel_path, data_only=True))


def resolve_faculty_codes(value: str, abbrev_map: dict) -> list:
    """Split a lab sheet's faculty cell into one display name per instructor.

    The lab workbook identifies staff by short name only ("AC"), while the
    lecture workbook writes "Ankush Chander (AC)". Left unresolved, every
    faculty-name query — schedules, locations, busy/free time — silently misses
    that person's labs.

        "AC"        -> ["Ankush Chander (AC)"]
        "AV/PK"     -> ["Ankit Vijayvargiya (AV)", "Pankaj Kumar (PK)"]
        "TF/TA"     -> ["TF/TA"]   # nothing resolvable — keep the cell whole
        ""          -> [""]        # still emit one record, instructor unknown

    One name per entry, never a joined string: `timetables.faculty_name` is
    matched by substring, so a compound value like "A (X) / B (Y)" would make
    A's own name ambiguous against it and resolve_faculty() could never narrow
    a query down — the caller then asks the user to disambiguate forever.

    Splitting only when at least one token resolves avoids inventing "TF" and
    "TA" as two separate people. Token matching is exact: "AC" and "AC1" are
    different people (Ankush Chander vs Arunava Chakravarty).
    """
    if not value:
        return [""]

    tokens = [t.strip() for t in value.split("/") if t.strip()]
    if not abbrev_map or not any(t in abbrev_map for t in tokens):
        return [value]

    return [abbrev_map.get(t, t) for t in tokens]


# ── Section Header Parsing ────────────────────────────────────────────────────
# Maps keywords found in Excel section headers to curriculum.json program names.
_HEADER_PROGRAM_MAP = {
    "BTech Core":            "B Tech (Institute Core)",
    "BTech (ICT and CS)":    "B Tech (ICT and CS)",
    "BTech (ICT-CS)":        "B Tech (ICT-CS)",
    "BTech (ICT, ICT-CS)":   "B Tech (ICT and ICT-CS)",
    "BTech (ICT &  CS)":     "B Tech (ICT and CS)",
    "BTech (ICT & CS)":      "B Tech (ICT and CS)",
    "BTech (CS)":            "B Tech (CS)",
    "BTech (MnC)":           "B Tech (MnC)",
    "BTech (MNC)":           "B Tech (MnC)",
    "BTech (EVD)":           "B Tech (EVD)",
    "BTech (ICT)":           "B Tech (Program Core)",
    "BTech":                 "B Tech (Institute Core)",
    "BS-MS (IT)":            "BS-MS (IT)",
    "BS-MS (DS & AI)":       "BS-MS (DS & AI)",
    "BS-MS (DS &amp; AI)":   "BS-MS (DS & AI)",
    # Deliberately unbalanced: real headers read "MTech (ICT , SS, ML, VES,
    # WCSP )Elective" and "MTech (ICT - SS , ML, VES)", so only the prefix is
    # stable. Do not "correct" it to "MTech (ICT)".
    "MTech (ICT":            "M Tech (ICT)",
    "MTech":                 "M Tech (ICT)",
    "MSc (IT)":              "MSc (IT)",
    "MSc (DS)":              "MSc (DS)",
    "MSc (AA)":              "MSc (AA)",
    "MDes (CD)":             "MDes (CD)",
    "MDes (IUxD)":           "MDes (IUxD)",
}

_ROMAN = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7, "VIII": 8}

# Sessions attributed to their section because the curriculum does not list the
# course under that program. Reset per parse; summarised at the end of the run.
_fallback_attributions = []

def _parse_section_header(header: str) -> tuple:
    """Extract (curriculum_program, semester_int) from an Excel section header.
    
    Examples:
        'MSc (IT) Core: SEMESTER III (2025 Batch)' → ('MSc (IT)', 3)
        'BTech Core: SEMESTER I (2026 Batch)'      → ('B Tech (Institute Core)', 1)
    Returns (None, None) if the header cannot be parsed.
    """
    # Extract semester number from "SEMESTER III" pattern
    semester = None
    sem_match = re.search(r"SEMESTER\s+(VIII|VII|VI|IV|V|III|II|I)\b", header, re.IGNORECASE)
    if sem_match:
        semester = _ROMAN.get(sem_match.group(1).upper())
    
    # Match program: try longest keys first to avoid partial matches
    program = None
    for key in sorted(_HEADER_PROGRAM_MAP, key=len, reverse=True):
        if key in header:
            program = _HEADER_PROGRAM_MAP[key]
            break
    
    return program, semester

def _filter_meta(meta_list: list, section_program: str, section_semester: int) -> list:
    """Filter curriculum entries to match the current section's program and semester.
    
    Strategy (best match first):
    1. Exact match on both program and semester → use it
    2. Match program only (semester is None in curriculum) → use it
    3. No filter matched → return all entries (original behaviour as fallback)
    """
    if not section_program or not meta_list:
        return meta_list
    
    # Exact match: both program and semester
    exact = [m for m in meta_list
             if m.get("program") == section_program
             and m.get("semester") == section_semester]
    if exact:
        return exact
    
    # Program-only match (covers electives where semester may be None)
    prog_only = [m for m in meta_list if m.get("program") == section_program]
    if prog_only:
        return prog_only
    
    # Fallback: The course is in curriculum, but doesn't map to this section's program.
    # Don't return all entries (causes bleeding into unrelated programs).
    # Instead, construct a single fallback entry for the current section,
    # borrowing the course_name from the first curriculum entry.
    #
    # This is a guess, and a wrong one is invisible in the data: the row simply
    # claims to belong to whichever section it appeared under. Record it, so a
    # reshaped workbook shows up as a rising count here rather than as silently
    # mislabelled rows in `timetables`. Expected to be non-zero — cross-program
    # electives legitimately take this path — so it is summarised per run rather
    # than warned per row, and the detail is kept at DEBUG.
    _fallback_attributions.append(
        (section_program, section_semester, sorted({m.get("program") for m in meta_list}))
    )
    logger.debug(
        f"No curriculum entry under section program {section_program!r} "
        f"(semester {section_semester}); curriculum lists this course under "
        f"{sorted({m.get('program') for m in meta_list})}"
    )
    fallback_name = meta_list[0].get("course_name", "") if meta_list else ""
    fallback_type = meta_list[0].get("course_type", "Unknown") if meta_list else "Unknown"
    return [{"program": section_program, "semester": section_semester,
             "course_name": fallback_name, "course_type": fallback_type}]

def _dedupe_records(records: list) -> list:
    """Collapse rows describing the same physical session for the same program.

    A session can be listed under more than one section of the lab workbook —
    IE494 appears under both "BTech (MnC) Elective: SEMESTER V" and
    "... SEMESTER VII". The VII section exact-matches one curriculum entry; the
    V section finds no exact match, falls back to program-only, and picks up
    both the semester-less and the semester-7 entry. The result is the same lab
    emitted two or three times.

    Where that happens, keep the row carrying a concrete semester — a curriculum
    entry with `semester: None` is the vaguer statement of the same fact.
    """
    # `semester` is stringified at record build time, so a missing one arrives
    # as "None" or "" rather than a falsy value.
    def has_semester(rec) -> bool:
        return rec["semester"] not in ("", "None")

    best = {}
    for rec in records:
        key = (rec["session_type"], rec["day"], rec["start"], rec["end"],
               rec["course_code"], rec["program"], rec["faculty"], rec["room"])
        current = best.get(key)
        if current is None or (not has_semester(current) and has_semester(rec)):
            best[key] = rec

    dropped = len(records) - len(best)
    if dropped:
        logger.info(f"Collapsed {dropped} duplicate lab session row(s).")
    return list(best.values())


def parse_excel(excel_path: str, curriculum: dict) -> list:
    """Parse the lecture workbook.

    Note: unlike the lab workbook, this sheet has no program section headers —
    column 0 holds only time slots, and each day column carries course/faculty/
    room. There is therefore nothing to pass to `_filter_meta`, and a lecture's
    programs can only come from fanning out over its curriculum entries. That
    asymmetry is inherent to the two sources, not an oversight: lecture rows
    keep category-style programs like "General Elective (Technical)" that lab
    rows no longer produce.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Abbreviation map: initials → full name
    abbrev_map = read_abbrev_map(wb)

    sheet_name = "Lecture (Update)" if "Lecture (Update)" in wb.sheetnames else wb.sheetnames[0]
    all_rows = list(wb[sheet_name].iter_rows(values_only=True))

    header_idx = None
    day_cols = {}
    for i, row in enumerate(all_rows):
        if any(str(v).strip() in DAYS for v in row if v):
            header_idx = i
            break
    if header_idx is None:
        logger.error("Could not find day header row in Excel.")
        sys.exit(1)
    for ci, v in enumerate(all_rows[header_idx]):
        if v and str(v).strip() in DAYS:
            day_cols[str(v).strip()] = ci

    records = []
    cur_time = None
    
    for row in all_rows[header_idx + 1:]:
        t = row[0] if len(row) > 0 else None
        if t and str(t).strip():
            cur_time = str(t).strip()
        if not cur_time:
            continue

        start, end = _resolve_slot(cur_time)

        for day in DAYS:
            if day not in day_cols:
                continue
            dc = day_cols[day]
            course  = row[dc]     if len(row) > dc     else None
            faculty = row[dc + 1] if len(row) > dc + 1 else None
            room    = row[dc + 2] if len(row) > dc + 2 else None

            if not course:
                continue
            cs = str(course).strip()
            fs = str(faculty).strip() if faculty else ""
            rs = str(room).strip()    if room    else ""

            if not cs or cs in ("-", "—", "Course"):
                continue

            faculty_full = abbrev_map.get(fs, f"{fs} (unresolved)" if fs else "TBA")

            # Look up metadata from curriculum
            # The excel has course codes. Curriculum JSON has them as keys.
            meta_list = curriculum.get(cs, [])
            if not meta_list:
                meta_list = [{}]  # fallback for unknown courses
                
            for meta in meta_list:
                for room_name in split_rooms(rs):
                    rec = {
                        "session_type": "Lecture",
                        "day": day,
                        "start": start,
                        "end": end,
                        "course_code": cs,
                        "course_name": meta.get("course_name", cs),
                        "course_type": meta.get("course_type", "Unknown"),
                        "program": meta.get("program", "Unknown Program"),
                        "semester": str(meta.get("semester", "")),
                        "faculty": faculty_full,
                        "room": room_name,
                    }
                    records.append(rec)

    return records

def parse_lab_excel(excel_path: str, curriculum: dict, abbrev_map: dict = None) -> list:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = wb.sheetnames[0]
    all_rows = list(wb[sheet_name].iter_rows(values_only=True))

    records = []
    
    header_idx = None
    for i, row in enumerate(all_rows):
        if any(str(v).strip() == "Time Slot" for v in row if v):
            header_idx = i
            break
            
    if header_idx is None:
        logger.warning(f"Could not find 'Time Slot' header in {excel_path}")
        return []

    day_cols = {}
    for ci, v in enumerate(all_rows[header_idx]):
        if v and str(v).strip() in DAYS:
            day_cols[str(v).strip()] = ci

    if not day_cols:
        day_cols = {"Monday": 1, "Tuesday": 2, "Wednesday": 3, "Thursday": 4, "Friday": 5}
        
    time_col, room_col, course_col, faculty_col = 0, 6, 7, 8
    current_program_header = "Unknown Program"
    section_program = None
    section_semester = None
    _fallback_attributions.clear()
    
    # Tutorial continuation rows: some rows have group data in day columns
    # but leave course/faculty blank, expecting them to carry forward.
    last_course = None
    last_faculty = None
    
    for row in all_rows[header_idx + 1:]:
        t = row[time_col] if len(row) > time_col else None
        
        if t and not re.search(r"\d{1,2}:\d{2}", str(t)):
            current_program_header = str(t).strip()
            section_program, section_semester = _parse_section_header(current_program_header)
            # Reset carry-forward on new program section
            last_course = None
            last_faculty = None
            continue
            
        if not t or not str(t).strip():
            continue
            
        start, end = _resolve_slot(t)
        if not start or not end:
            continue
            
        room_raw = str(row[room_col]).strip() if len(row) > room_col and row[room_col] else ""
        course_raw = str(row[course_col]).strip() if len(row) > course_col and row[course_col] else ""
        faculty_raw = str(row[faculty_col]).strip() if len(row) > faculty_col and row[faculty_col] else ""
        
        # Carry-forward: update trackers when present, inherit when blank
        if course_raw and course_raw not in ("-", "—"):
            last_course = course_raw
        if faculty_raw:
            last_faculty = faculty_raw
        
        course = course_raw or last_course or ""
        faculty_cell = faculty_raw or last_faculty or ""
        # Lab sheets carry short names only — resolve to the same display form
        # the lecture sheet writes, so both are reachable by one name lookup.
        # One record per instructor: faculty_name must name exactly one person.
        faculty_names = resolve_faculty_codes(faculty_cell, abbrev_map)
        
        if not course or course in ("-", "—"):
            continue

        # New architecture: lookup (course, program, semester)
        meta_list = curriculum.get(course, [])
        if meta_list:
            meta_list = _filter_meta(meta_list, section_program, section_semester)
        if not meta_list:
            meta_list = [{"program": current_program_header, "semester": section_semester}]
            
        for day in DAYS:
            if day not in day_cols:
                continue
            dc = day_cols[day]
            group_val = row[dc] if len(row) > dc else None
            if group_val and str(group_val).strip():
                gv = str(group_val).strip()
                session_type = f"Tutorial ({gv})" if "tut" in gv.lower() else f"Lab ({gv})"
                
                for meta in meta_list:
                    for room_name in split_rooms(room_raw):
                        for faculty_name in faculty_names:
                            records.append({
                                "session_type": session_type,
                                "day": day,
                                "start": start,
                                "end": end,
                                "course_code": course,
                                "course_name": meta.get("course_name", course),
                                "course_type": meta.get("course_type", "Unknown"),
                                "program": meta.get("program", current_program_header),
                                "semester": str(meta.get("semester", "")),
                                "faculty": faculty_name,
                                "room": room_name,
                            })

    if _fallback_attributions:
        by_section = Counter((p, s) for p, s, _ in _fallback_attributions)
        logger.warning(
            f"{len(_fallback_attributions)} lab session(s) attributed to the section "
            f"they appeared under because the curriculum does not list the course "
            f"for that program. Mostly cross-program electives; a jump here after a "
            f"workbook change means section headers stopped matching. "
            f"By section: {dict(by_section)}. Run at DEBUG for per-course detail."
        )

    return _dedupe_records(records)


def main():
    try:
        excel_path = find_excel_file()
        logger.info(f"Parsing active timetable from: {excel_path}")
        
        curriculum = load_curriculum()
        logger.info(f"Loaded {len(curriculum)} courses from curriculum.json")
        
        records = parse_excel(excel_path, curriculum)
        
        lab_excel_path = find_lab_excel_file()
        if lab_excel_path:
            logger.info(f"Parsing lab timetable from: {lab_excel_path}")
            # The abbreviations sheet lives in the lecture workbook, so the lab
            # parser has to be handed it explicitly.
            abbrev_map = load_abbrev_map(excel_path)
            logger.info(f"Loaded {len(abbrev_map)} faculty short names for lab resolution.")
            lab_records = parse_lab_excel(lab_excel_path, curriculum, abbrev_map)
            records.extend(lab_records)
        else:
            logger.info("No lab timetable file found. Skipping labs.")

        logger.info(f"Parsed {len(records)} timetable records.")

        with db_connection() as conn:
            with conn.cursor() as cur:
                # TRUNCATE existing timetables to keep only active timetable dataset
                cur.execute("TRUNCATE TABLE timetables;")
                logger.info("Cleared existing timetables table.")

                INSERT_SQL = """
                    INSERT INTO timetables
                        (session_type, course_code, course_name, course_type, program, semester, faculty_name,
                         day_of_week, start_time, end_time, room)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                counts = {}
                for rec in records:
                    prog = rec["program"]
                    cur.execute(INSERT_SQL, (
                        rec["session_type"],
                        rec["course_code"],
                        rec["course_name"],
                        rec["course_type"],
                        rec["program"],
                        rec["semester"] if rec["semester"] else None,
                        rec["faculty"],
                        rec["day"],
                        rec["start"],
                        rec["end"],
                        rec["room"]
                    ))
                    counts[prog] = counts.get(prog, 0) + 1

                # Auto-vivify any rooms found in the timetable that are missing from the venues table.
                # This ensures every room in the timetable has a basic venue record (and its official POC assigned).
                from core import config
                cur.execute("""
                    INSERT INTO venues (venue_id, capacity, venue_type, booking_poc)
                    SELECT DISTINCT room,
                           1,
                           CASE WHEN room ILIKE 'CEP%%' THEN 'room'
                                WHEN room ILIKE '%%LT%%' THEN 'lt'
                                WHEN room ILIKE '%%LAB%%' THEN 'lab'
                                ELSE 'other' END,
                           CASE WHEN room ILIKE 'CEP%%' THEN %s
                                WHEN room ILIKE '%%LT%%' OR room ILIKE '%%LAB%%' THEN %s
                                ELSE NULL END
                    FROM timetables
                    WHERE room IS NOT NULL AND room <> ''
                    ON CONFLICT (venue_id) DO NOTHING
                """, (config.CEP_BOOKING_POC, config.LAB_LT_BOOKING_POC))
                
            conn.commit()

        logger.info("=== Timetable Seeded Successfully ===")
        for prog, n in sorted(counts.items()):
            logger.info(f"  {prog}: {n} rows")
        logger.info(f"Total inserted: {sum(counts.values())} records.")

    except Exception as e:
        logger.error(f"Failed to seed timetable data: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
