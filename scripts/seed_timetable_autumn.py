"""
Seed Autumn 2026-27 Timetable
==============================
Parses Lecture_TT_Autumn2026-27_v9.xlsx and inserts all programs into
the timetables table. Safe to re-run — deletes existing Autumn 2026-27
rows first, then re-inserts fresh.

Run AFTER seed_timetable.py (which seeds the Spring/base term data).

Usage:
    python scripts/seed_timetable_autumn.py
"""
import re
import os
import sys
import psycopg2

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from core import config

logger = config.get_logger("scripts.seed_timetable_autumn")

try:
    import openpyxl
except ImportError:
    logger.error("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "Lecture_TT_Autumn2026-27_v9.xlsx",
)
TERM = "Autumn 2026-27"
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# Batch label in Excel → clean program name
LABEL_MAP = {
    "Btech 1st Yr"        : "BTech Sem-I",
    "Btech 2nd Year"      : "BTech Sem-III",
    "Btech 2nd Yr"        : "BTech Sem-III",
    "Btech 3r Yr"         : "BTech Sem-V",
    "Btech 3rd Yr"        : "BTech Sem-V",
    "Btech 3rd Yr (Core)" : "BTech Sem-V (Core)",
    "Btech Core"          : "BTech Core (Elective)",
    "Elective"            : "Elective (Open)",
    "MSc (AA)"            : "MSc Sem-I (AA)",
    "MSC (AA)"            : "MSc Sem-I (AA)",
    "MSc (DS) Core"       : "__MSCDS__",   # disambiguated by course code
    "MSc DS (Core)"       : "__MSCDS__",
    "MSc (IT) Core"       : "__MSCIT__",   # disambiguated by room number
    "BS-MS (IT)"          : "BS-MS (IT)",
    "BS-MS (DS + AI)"     : "BS-MS (DS + AI)",
    "Mtech (Core)"        : "MTech Sem-I",
}

# MSc DS: course code → semester (from academic requirements PDF)
MSCDS_SEM1_CODES = {"DS601", "DS602", "DS603", "DS604", "DS605"}
MSCDS_SEM3_CODES = {"DS635", "DS636", "DS639"}

# MSc IT: room → semester
MSCIT_ROOM_MAP = {
    "CEP207": "MSc Sem-I (IT)",
    "CEP209": "MSc Sem-III (IT)",
}


def _normalise_room(room: str) -> str:
    return room.upper().replace(" ", "").replace("-", "")


def _parse_time(t) -> str | None:
    if not t:
        return None
    m = re.match(r"(\d{1,2}):(\d{2})", str(t))
    return f"{int(m.group(1)):02d}:{m.group(2)}:00" if m else None


def _resolve_slot(slot) -> tuple:
    if not slot:
        return None, None
    m = re.match(r"(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})", str(slot))
    if m:
        return _parse_time(m.group(1)), _parse_time(m.group(2))
    return None, None


def parse_excel(excel_path: str) -> tuple[list, list]:
    """
    Returns (records, needs_review).
    Each record: {program, day, start, end, course, faculty, room}
    """
    if not os.path.exists(excel_path):
        logger.error(f"Excel file not found: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Abbreviation map: initials → full name
    abbrev_map = {}
    for row in wb["FacultyNameAbbreviations"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            abbrev_map[str(row[1]).strip()] = str(row[0]).strip()

    all_rows = list(wb["Lecture (Update)"].iter_rows(values_only=True))

    # Find header row
    header_idx = None
    day_cols = {}
    for i, row in enumerate(all_rows):
        if any(str(v).strip() in DAYS for v in row if v):
            header_idx = i
            break
    if header_idx is None:
        logger.error("Could not find day header row in Excel.")
        sys.exit(1)
    for ci, v in enumerate(all_rows[header_idx]):
        if v and str(v).strip() in DAYS:
            day_cols[str(v).strip()] = ci

    records = []
    needs_review = []
    cur_time = cur_batch = None
    unknown_labels = set()

    for row in all_rows[header_idx + 1:]:
        t = row[0] if len(row) > 0 else None
        b = row[2] if len(row) > 2 else None
        if t and str(t).strip():
            cur_time = str(t).strip()
        if b and str(b).strip():
            cur_batch = str(b).strip()
        if not cur_time or not cur_batch:
            continue

        mapped = LABEL_MAP.get(cur_batch)
        if mapped is None:
            unknown_labels.add(cur_batch)
            continue

        start, end = _resolve_slot(cur_time)

        for day in DAYS:
            if day not in day_cols:
                continue
            dc = day_cols[day]
            course  = row[dc]     if len(row) > dc     else None
            faculty = row[dc + 1] if len(row) > dc + 1 else None
            room    = row[dc + 2] if len(row) > dc + 2 else None

            if not course:
                continue
            cs = str(course).strip()
            fs = str(faculty).strip() if faculty else ""
            rs = str(room).strip()    if room    else ""

            # Skip blank/artifact rows
            if not cs or cs in ("-", "—", "Course"):
                continue

            faculty_full = abbrev_map.get(fs, f"{fs} (unresolved)" if fs else "TBA")

            rec = {
                "day": day, "start": start, "end": end,
                "course": cs, "faculty": faculty_full, "room": rs,
            }

            # ── MSc IT: room-based disambiguation ──────────────────────────
            if mapped == "__MSCIT__":
                r_clean = _normalise_room(rs)
                program = None
                for key, prog in MSCIT_ROOM_MAP.items():
                    if key in r_clean:
                        program = prog
                        break
                if program:
                    rec["program"] = program
                    records.append(rec)
                else:
                    rec["program"] = "MSc (IT) Core"
                    rec["review_reason"] = f"unrecognised room '{rs}' (expected CEP-207 or CEP-209)"
                    needs_review.append(rec)

            # ── MSc DS: course-code-based disambiguation ────────────────────
            elif mapped == "__MSCDS__":
                if cs in MSCDS_SEM1_CODES:
                    rec["program"] = "MSc Sem-I (DS)"
                    records.append(rec)
                elif cs in MSCDS_SEM3_CODES or cs.startswith("DS63"):
                    rec["program"] = "MSc Sem-III (DS)"
                    records.append(rec)
                else:
                    rec["program"] = "MSc (DS) Core"
                    rec["review_reason"] = f"course '{cs}' not in Sem-I or Sem-III DS list"
                    needs_review.append(rec)

            # ── All other programs ──────────────────────────────────────────
            else:
                rec["program"] = mapped
                records.append(rec)

    if unknown_labels:
        logger.warning(f"Unknown batch labels (skipped): {unknown_labels}")

    return records, needs_review


def seed(conn, records: list, needs_review: list):
    cur = conn.cursor()

    # Delete existing Autumn 2026-27 rows (idempotent re-run)
    autumn_programs = list({r["program"] for r in records})
    cur.execute(
        "DELETE FROM timetables WHERE batch_group = ANY(%s)",
        (autumn_programs,),
    )
    deleted = cur.rowcount
    logger.info(f"Cleared {deleted} existing Autumn 2026-27 rows.")

    INSERT_SQL = """
        INSERT INTO timetables
            (session_type, course_code, course_name, faculty_name,
             day_of_week, start_time, end_time, location, batch_group)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    counts = {}
    for rec in records:
        prog = rec["program"]
        cur.execute(INSERT_SQL, (
            "Lecture",
            rec["course"], rec["course"],
            rec["faculty"],
            rec["day"],
            rec["start"], rec["end"],
            rec["room"],
            prog,
        ))
        counts[prog] = counts.get(prog, 0) + 1

    conn.commit()

    logger.info(f"=== Autumn 2026-27 Timetable Seeded ===")
    for prog, n in sorted(counts.items()):
        logger.info(f"  {prog}: {n} rows")
    logger.info(f"  Total: {sum(counts.values())} rows inserted")

    if needs_review:
        logger.warning(f"  {len(needs_review)} rows need manual review:")
        for r in needs_review:
            logger.warning(f"    {r['day']} | {r['course']} | room={r['room']} | {r.get('review_reason','')}")

    cur.close()


if __name__ == "__main__":
    from core.database import db_connection

    logger.info("=== Autumn 2026-27 Timetable Seeder ===")
    logger.info(f"Parsing: {EXCEL_PATH}")

    records, needs_review = parse_excel(EXCEL_PATH)
    logger.info(f"Parsed {len(records)} records, {len(needs_review)} need review.")

    with db_connection() as conn:
        seed(conn, records, needs_review)

    logger.info("Done.")
